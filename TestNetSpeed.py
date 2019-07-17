#!/usr/bin/python
import speedtest, openpyxl, datetime, requests, calendar
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet import dimensions
import time

def testSpeed():
    s = speedtest.Speedtest()
    s.get_servers()
    s.get_best_server()
    s.upload()
    s.download()
    res = s.results.dict()
    return res['download'], res['upload'], res['ping']

def connectedToInternet():
    fail = 0
    URLS = {
        "Google": "https://www.google.com",
        "Bing" : "https://www.google.com",
        "Yahoo" : "https://www.yahoo.com",
        "Amazon" :"https://www.amazon.com"
    }
    for url in URLS: 
        try:
            requests.get(URLS[url])
            return True
        except:
            fail = fail + 1
    if fail == len(URLS):
        return False

    

def getWorkBook(year):
    try:
        wb = openpyxl.load_workbook('Documents/Speedtests/{}.xlsx'.format(year))

    except:
        wb = openpyxl.Workbook()
        wb.save('Documents/Speedtests/{}.xlsx'.format(year))
        #wb.save('Documents/Tests/{}.xlsx'.format(year))
        try:
            del wb['Sheet']
        except:
            pass
    return wb

def getSheet(wb):
    week = int(getNow().strftime('%V'))
    if int(getNow().strftime('%w')) == 0:
        week = week + 1
    try:
        sheet = wb['Week {}'.format(week)]
    except:
        sheet = wb.create_sheet('Week {}'.format(week))
    return sheet

def getCells(sheet):   
    columns = sheet.max_column
    weekday = getNow().strftime('%A') 
    if columns == 1:
        initializeWeek(sheet)
    recordSpeed(sheet)
    

def initializeWeek(sheet):
    for i in range (2, ((7*3)+2)):

        if (i+1)%3 == 0:
            sheet.merge_cells('{}1:{}1'.format(get_column_letter(i), get_column_letter(i+2)))
            sheet.cell(row = 1, column = i).value = calendar.day_name[(int(i/3)+6)%7]
            sheet.cell(row = 1, column = i).alignment = Alignment(horizontal='center')
            sheet.cell(row = 2, column = i). value = 'Download (Kb/s)'
        if i%3==0:
            sheet.cell(row = 2, column = i). value = 'Upload (Kb/s)'

        if (i+2)%3==0:
            sheet.cell(row = 2, column = i). value = 'Ping'       

    for i in range (1, 24):
        t = datetime.time(i, 0)
        
        AM = t.strftime('%p')
        sheet.cell(row = (i + 2), column = 1).value = t.strftime('%I:%m %p')
    
            
def getNow():
    return datetime.datetime.now()


def recordSpeed(ws):
    
    hour = int(getNow().strftime('%H')) + 2 + 1
    intday = (int(getNow().strftime('%w')) + 1)%7
    day = (((intday)%7)*3) - 1
    if connectedToInternet():
    #if False:
        down, up, ping = testSpeed()
        ws.cell(row = hour, column = day).value = down/1024
        ws.cell(row = hour, column = day + 1).value = up/1024
        ws.cell(row = hour, column = day + 2).value = ping
    else:
        ws.cell(row = hour, column = day).value = 0
        ws.cell(row = hour, column = day + 1).value = 0
        ws.cell(row = hour, column = day + 2).value = 0

def main():
    year = getNow().year
    wb = getWorkBook(year)
    ws = getSheet(wb)
    getCells(ws)
    
    saveSucceeded = False
    while saveSucceeded == False:
        try:
            print('Saving')
            wb.save('Documents/Speedtests/{}.xlsx'.format(year))
            saveSucceeded = True
        except: 
            print('Could not save file\n')
            time.sleep(300)
            

if __name__ == '__main__':
    main()
